from openpyxl import load_workbook
import json
import os
"""
Convierte el protocolo excel y lo convierte en formato .SMVA

"""
def excel_to_smva(PATH):

    #wb = load_workbook(r"_TEMPS_\test_archivo_smva3.xlsx")
    wb = load_workbook(PATH)
    ws = wb.active 
    ultima_fila = ws.max_row
    fila_num = 1
    coord = {}
    ruta, archivo = os.path.split(PATH) #Esto me sirve para separar y guardarlo en el mismo lugar
    nombre_archivo, _ = os.path.splitext(archivo) #esto me sirve para agregar el nuevo nombre
    nombre_archivo = f"{nombre_archivo}.SMVA"
    SAVE_PATH = os.path.join(ruta,nombre_archivo)
    # Estado de si se encontraron los encabezados
    encontrados = {
        "bloque": False,
        "descripcion": False,
        "central": False,
        "minimo": False,
        "maximo": False,
        "tiempo": False,
        "validacion": False,
        "tiporesp": False,
        "comparacion": False,
        "comando": False,
        "respuesta": False,
        "item":False,
        "factor":False,
        "unidad":False
    }

    while fila_num < 20:
        for celda in ws[fila_num]:
            if celda.value:
                texto = str(celda.value).strip().lower()

                if "bloque" in texto and not encontrados["bloque"]:
                    #print(f"'bloque' encontrado en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Nombre"] = [celda.column_letter, fila_num]
                    encontrados["bloque"] = True

                elif "descripción para operador" in texto and not encontrados["descripcion"]:
                    #print(f"'descripcion' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["descripcion"] = [celda.column_letter, fila_num]
                    encontrados["descripcion"] = True

                elif "conv" in texto and not encontrados["factor"]:
                    #print(f"'descripcion' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["FactorConversion"] = [celda.column_letter, fila_num]
                    encontrados["factor"] = True
                elif "unid" in texto and not encontrados["unidad"]:
                    #print(f"'descripcion' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Unidad"] = [celda.column_letter, fila_num]
                    encontrados["unidad"] = True
                elif "item" in texto and not encontrados["item"]:
                    #print(f"'descripcion' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Tipo_Item"] = [celda.column_letter, fila_num]
                    encontrados["item"] = True
                elif "central" in texto and not encontrados["central"]:
                    #print(f"'central' encontrado en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["ResultadoTipico"] = [celda.column_letter, fila_num]
                    encontrados["central"] = True

                elif "min" in texto and not encontrados["minimo"]:
                    #print(f"'minimo' encontrado en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["ResultadoMinimo"] = [celda.column_letter, fila_num]
                    encontrados["minimo"] = True

                elif "max" in texto and not encontrados["maximo"] or "máx" in texto and not encontrados["maximo"]:
                    #print(f"'maximo' encontrado en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["ResultadoMaximo"] = [celda.column_letter, fila_num]
                    encontrados["maximo"] = True

                elif "tiempo" in texto and not encontrados["tiempo"]:
                    #print(f"'tiempo' encontrado en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Tiempo_Medicion"] = [celda.column_letter, fila_num]
                    encontrados["tiempo"] = True

                elif "valida" in texto and not encontrados["validacion"]:
                    #print(f"'validacion' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Validacion"] = [celda.column_letter, fila_num]
                    encontrados["validacion"] = True

                elif "respuesta" in texto and not encontrados["tiporesp"]:
                    #print(f"'tipo respuesta' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Tipo_Respuesta"] = [celda.column_letter, fila_num]
                    encontrados["tiporesp"] = True

                elif "comparacion" in texto and not encontrados["comparacion"]:
                    #print(f"'comparacion' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Tipo"] = [celda.column_letter, fila_num]
                    encontrados["comparacion"] = True

                elif "comando" in texto and not encontrados["comando"]:
                    #print(f"'comando' encontrado en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Comandos"] = [celda.column_letter, fila_num]
                    encontrados["comando"] = True

                elif "correct" in texto and not encontrados["respuesta"]:
                    #print(f"'respuesta' encontrada en la columna: {celda.column_letter}, fila: {fila_num}")
                    coord["Respuesta_Correcta"] = [celda.column_letter, fila_num]
                    encontrados["respuesta"] = True

        fila_num += 1
        if all(encontrados.values()):
            break

    #print("\nCoordenadas encontradas:")
    #print(coord)
    bloques = []
    bloque_actual = None
    pasos_actuales = []
    N=-1
    n_secuencia=0
    fila_inicio = min(c[1] for c in coord.values()) + 1
    fila_inicio = coord["Nombre"][1] #(A,5)
    for fila in range(fila_inicio+1, ultima_fila + 1):
        celda_bloque = coord["Nombre"][0] + str(fila)
        valor_bloque = ws[celda_bloque].value

        if valor_bloque:
            # Si ya había un bloque en curso, lo guardamos antes de empezar el nuevo
            if bloque_actual:
                n_secuencia=0
                N+=1
                bloques.append({
                    "ProtocoloID":N,
                    "Nombre": bloque_actual,
                    "Pasos": pasos_actuales,
                    "ordenSecuencia": N,
                    "Resultado": "PASA",
                    "Operador": ""
                })
            bloque_actual = str(valor_bloque).strip()
            pasos_actuales = []

        if not bloque_actual:
            continue

        paso = {}
        for clave in coord:
            if clave == "Nombre":
                continue
            col, _ = coord[clave]
            valor = ws[f"{col}{fila}"].value
            if clave =="descripcion":
                clave ="Nombre"
            elif clave=="ResultadoTipico" or clave =="ResultadoMinimo" or clave =="ResultadoMaximo" or clave=="FactorConversion":
                if valor == None:
                    valor = 0.0
                else:
                    valor = float(valor)
            elif clave=="Tipo" or clave=="Unidad":
                if valor ==None:
                    valor =""
            elif clave =="Comandos":
                if valor ==None:
                    valor =""
            elif clave == "Tipo_Item":
                if valor =="IM":
                    valor = "IngresoManual"
                elif valor =="M":
                    valor ="Medicion"
                elif valor =="PI":
                    valor = "ProgramacionDeInstrumento"
                elif valor =="VR":
                    valor ="Verificacion"
                else:
                    valor ="Calibracion"
            elif clave =="Respuesta_Correcta":
                if valor ==None:
                    valor =""
            elif clave =="Tiempo_Medicion":
                try:
                    valor = float(valor)
                except:
                    valor = 0.2
            elif clave =="Tipo_Respuesta":
                try:
                    if "text" in valor.lower():
                        valor = "TEXTO"
                    else:
                        valor ="NUMERICO"
                except:
                    valor = "TEXTO"
            paso[clave] = valor
        paso["OrdenDeSecuencia"]=str(n_secuencia)
        n_secuencia+=1
        paso["Estado"]=""
        paso["Resultado"]=""
        paso["CriterioPass"]=""
        paso["TimeStamp"]=""
        paso["protocolo_idprotocolo"]=fila
        paso["protocolo_protocolos_idProtocolos"]=fila+1
        paso["mediciones_idmediciones"]=fila+2
        paso["id_paso"]=n_secuencia+3
        pasos_actuales.append(paso)

    # No olvides guardar el último bloque procesado
    if bloque_actual and pasos_actuales:
        N+=1
        bloques.append({
            "ProtocoloID":N,
            "Nombre": bloque_actual,
            "Pasos": pasos_actuales,
            "ordenSecuencia": N,
            "Resultado": "PASA",
            "Operador": ""
        })




    with open(SAVE_PATH,"w") as file:
        json.dump(bloques,file,indent=2,ensure_ascii=True)
        
#excel_to_smva(PATH=r"C:\Users\juanc\Desktop\SISTEMA MEDICIONES\_TEMPS_\test_archivo_smva3.xlsx")


def load_smva_file(PATH,basedato=None): #La base de dato para controlar los saltos
    
    with open(PATH,"r") as file:
        PROTOCOLO_SMVA = json.load(file)
    #Carga el archivo en el temporal
    with open(r"_TEMPS_\protocolo_a_ejecutar.json","w",encoding='utf-8') as file:
        json.dump(PROTOCOLO_SMVA,file,ensure_ascii=False, indent=4)
    #Cargo los saltos condicionales en caso de haberlos
    if basedato!=None:
        for i, bloque in enumerate(PROTOCOLO_SMVA):
            for j, paso in enumerate(bloque["Pasos"]):
                comandos = paso.get("Comandos", "")
                if "ETIQ:" in comandos:
                    partes = comandos.split('ETIQ:') #cAMBIO ESTA SINTAXIS PARA SER MAS EFICIENTE
                    if len(partes) > 1:
                        etiqueta = partes[1]
                        if ";" in etiqueta:
                            etiqueta = etiqueta.split(";")[0]
                        try:
                            etiqueta = etiqueta.split('"')[1] #aGREGO QUE LUEGO DEL ETIQ BUSQUE LO SIGUIENTE A LAS COMILLAS
                        except:
                            pass
                        basedato.SALTOS_CONDICIONALES[etiqueta.strip()] = {"i": i, "j": j} #VARIABLE QUE CONTROLA LOS SALTOS CONDICIONALES

#load_smva_file(PATH=r"C:\Users\juanc\Desktop\SISTEMA MEDICIONES\_TEMPS_\test_archivo_smva3.SMVA")